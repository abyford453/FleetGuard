import csv
import io
import json
from datetime import timedelta
from decimal import Decimal

from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.db.models.functions import TruncDate, TruncMonth, Coalesce
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from apps.fleet.models import Vehicle
from apps.inspections.models import Inspection, InspectionAlert
from apps.documents.models import VehicleDocument
from apps.fuel.models import FuelLog
from apps.fuel.alerts import vehicles_missing_fuel_logs, odometer_regressions


def _vehicle_label(v: Vehicle) -> str:
    label = v.unit_number or v.plate or "Vehicle"
    mm = f"{v.make} {v.model}".strip()
    if mm:
        return f"{label} ({mm})"
    return label


def _xlsx_response(wb: Workbook, filename: str) -> HttpResponse:
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def _autosize_columns(ws):
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            v = str(v)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[get_column_letter(col)].width = min(max(12, max_len + 2), 55)


def _write_sheet(ws, title: str, headers: list[str], rows: list[list]):
    ws.title = title
    ws.append(headers)

    header_font = Font(bold=True)
    for i in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=i)
        c.font = header_font
        c.alignment = Alignment(vertical="center")

    for r in rows:
        ws.append(r)

    _autosize_columns(ws)


def _build_report_context(request):
    tenant = request.tenant
    today = timezone.localdate()

    # KPI Snapshot
    vehicle_count = Vehicle.objects.filter(tenant=tenant).count()

    open_alerts = (
        InspectionAlert.objects
        .filter(tenant=tenant)
        .exclude(status=InspectionAlert.STATUS_CLOSED)
        .count()
    )

    overdue_inspections = (
        Inspection.objects
        .filter(tenant=tenant, due_date__isnull=False, due_date__lt=today)
        .exclude(status=Inspection.STATUS_COMPLETED)
        .count()
    )

    due_soon_inspections = (
        Inspection.objects
        .filter(
            tenant=tenant,
            due_date__isnull=False,
            due_date__gte=today,
            due_date__lte=today + timedelta(days=7),
        )
        .exclude(status=Inspection.STATUS_COMPLETED)
        .count()
    )

    expired_docs = (
        VehicleDocument.objects
        .filter(tenant=tenant, expires_on__isnull=False, expires_on__lt=today)
        .count()
    )
    expiring_docs = (
        VehicleDocument.objects
        .filter(
            tenant=tenant,
            expires_on__isnull=False,
            expires_on__gte=today,
            expires_on__lte=today + timedelta(days=30),
        )
        .count()
    )

    fuel_stale_count = len(vehicles_missing_fuel_logs(tenant, days=30))
    fuel_odo_alert_count = len(odometer_regressions(tenant))

    spend_30 = (
        FuelLog.objects
        .filter(tenant=tenant, fuel_date__gte=today - timedelta(days=30))
        .aggregate(total=Coalesce(Sum("cost"), Decimal("0.00")))["total"]
    )

    # Chart data
    start_30 = today - timedelta(days=30)
    daily = (
        FuelLog.objects
        .filter(tenant=tenant, fuel_date__gte=start_30)
        .exclude(cost__isnull=True)
        .values("fuel_date")
        .annotate(total=Coalesce(Sum("cost"), Decimal("0.00")))
        .order_by("fuel_date")
    )
    daily_labels = [row["fuel_date"].strftime("%Y-%m-%d") for row in daily]
    daily_values = [float(row["total"]) for row in daily]

    start_12m = today - timedelta(days=365)
    monthly = (
        FuelLog.objects
        .filter(tenant=tenant, fuel_date__gte=start_12m)
        .exclude(cost__isnull=True)
        .annotate(m=TruncMonth("fuel_date"))
        .values("m")
        .annotate(total=Coalesce(Sum("cost"), Decimal("0.00")))
        .order_by("m")
    )
    monthly_labels = [row["m"].strftime("%Y-%m") for row in monthly]
    monthly_values = [float(row["total"]) for row in monthly]

    start_90 = today - timedelta(days=90)
    top = (
        FuelLog.objects
        .filter(tenant=tenant, fuel_date__gte=start_90)
        .exclude(cost__isnull=True)
        .values("vehicle_id")
        .annotate(total=Coalesce(Sum("cost"), Decimal("0.00")))
        .order_by("-total")[:8]
    )
    vehicle_map = {v.id: _vehicle_label(v) for v in Vehicle.objects.filter(tenant=tenant)}
    top_rows = [(vehicle_map.get(row["vehicle_id"], f"Vehicle #{row['vehicle_id']}"), float(row["total"])) for row in top]
    top_labels = [r[0] for r in top_rows]
    top_values = [r[1] for r in top_rows]

    return {
        "today": today,
        "vehicle_count": vehicle_count,
        "open_alerts": open_alerts,
        "overdue_inspections": overdue_inspections,
        "due_soon_inspections": due_soon_inspections,
        "expired_docs": expired_docs,
        "expiring_docs": expiring_docs,
        "fuel_stale_count": fuel_stale_count,
        "fuel_odo_alert_count": fuel_odo_alert_count,
        "spend_30": spend_30,
        "daily_labels_json": json.dumps(daily_labels),
        "daily_values_json": json.dumps(daily_values),
        "monthly_labels_json": json.dumps(monthly_labels),
        "monthly_values_json": json.dumps(monthly_values),
        "top_labels_json": json.dumps(top_labels),
        "top_values_json": json.dumps(top_values),
        "top_rows": top_rows,
    }


@login_required
def index(request):
    return render(request, "reports/index.html", _build_report_context(request))


@login_required
def print_report(request):
    # Print-friendly layout (use browser Print → Save as PDF)
    return render(request, "reports/print.html", _build_report_context(request))


# ---------------- CSV EXPORTS ----------------

@login_required
def export_fuel_csv(request):
    tenant = request.tenant
    today = timezone.localdate()
    start = today - timedelta(days=365)

    qs = (
        FuelLog.objects
        .filter(tenant=tenant, fuel_date__gte=start)
        .select_related("vehicle")
        .order_by("-fuel_date", "-created_at")
    )

    resp = HttpResponse(content_type="text/csv")
    resp["Content-Disposition"] = 'attachment; filename="fuel_logs_last_12_months.csv"'
    w = csv.writer(resp)
    w.writerow(["fuel_date", "vehicle", "odometer", "gallons", "cost", "vendor", "fuel_type", "notes"])

    for r in qs:
        w.writerow([r.fuel_date, _vehicle_label(r.vehicle), r.odometer or "", r.gallons, r.cost or "", r.vendor, r.fuel_type, r.notes])
    return resp


@login_required
def export_inspections_csv(request):
    tenant = request.tenant
    qs = Inspection.objects.filter(tenant=tenant).select_related("vehicle").order_by("-created_at")

    resp = HttpResponse(content_type="text/csv")
    resp["Content-Disposition"] = 'attachment; filename="inspections.csv"'
    w = csv.writer(resp)
    w.writerow(["created_at", "vehicle", "inspection_type", "status", "due_date", "performed_on", "notes"])

    for i in qs:
        v = getattr(i, "vehicle", None)
        w.writerow([
            getattr(i, "created_at", "").strftime("%Y-%m-%d %H:%M") if getattr(i, "created_at", None) else "",
            _vehicle_label(v) if v else "",
            getattr(i, "inspection_type", ""),
            getattr(i, "status", ""),
            getattr(i, "due_date", "") or "",
            getattr(i, "performed_on", "") or "",
            getattr(i, "notes", ""),
        ])
    return resp


@login_required
def export_documents_csv(request):
    tenant = request.tenant
    qs = VehicleDocument.objects.filter(tenant=tenant).select_related("vehicle").order_by("-uploaded_at")

    resp = HttpResponse(content_type="text/csv")
    resp["Content-Disposition"] = 'attachment; filename="documents.csv"'
    w = csv.writer(resp)
    w.writerow(["uploaded_at", "vehicle", "doc_type", "title", "expires_on", "file"])

    for d in qs:
        w.writerow([
            d.uploaded_at.strftime("%Y-%m-%d %H:%M") if getattr(d, "uploaded_at", None) else "",
            _vehicle_label(d.vehicle),
            d.doc_type,
            d.title,
            d.expires_on or "",
            getattr(d.file, "url", ""),
        ])
    return resp


@login_required
def export_inspection_alerts_csv(request):
    tenant = request.tenant
    qs = InspectionAlert.objects.filter(tenant=tenant).select_related("vehicle").order_by("-created_at")

    resp = HttpResponse(content_type="text/csv")
    resp["Content-Disposition"] = 'attachment; filename="inspection_alerts.csv"'
    w = csv.writer(resp)
    w.writerow(["created_at", "vehicle", "severity", "status", "title", "detail"])

    for a in qs:
        v = getattr(a, "vehicle", None)
        w.writerow([
            a.created_at.strftime("%Y-%m-%d %H:%M") if getattr(a, "created_at", None) else "",
            _vehicle_label(v) if v else "",
            getattr(a, "severity", ""),
            getattr(a, "status", ""),
            getattr(a, "title", ""),
            getattr(a, "detail", ""),
        ])
    return resp


# ---------------- EXCEL EXPORTS ----------------

@login_required
def export_fuel_xlsx(request):
    tenant = request.tenant
    today = timezone.localdate()
    start = today - timedelta(days=365)

    qs = (
        FuelLog.objects
        .filter(tenant=tenant, fuel_date__gte=start)
        .select_related("vehicle")
        .order_by("-fuel_date", "-created_at")
    )

    wb = Workbook()
    ws = wb.active
    rows = []
    for r in qs:
        rows.append([
            r.fuel_date,
            _vehicle_label(r.vehicle),
            r.odometer or "",
            float(r.gallons),
            float(r.cost) if r.cost is not None else "",
            r.vendor,
            r.fuel_type,
            r.notes,
        ])

    _write_sheet(ws, "Fuel Logs", ["Fuel Date", "Vehicle", "Odometer", "Gallons", "Cost", "Vendor", "Fuel Type", "Notes"], rows)
    return _xlsx_response(wb, "fuel_logs_last_12_months.xlsx")


@login_required
def export_inspections_xlsx(request):
    tenant = request.tenant
    qs = Inspection.objects.filter(tenant=tenant).select_related("vehicle").order_by("-created_at")

    wb = Workbook()
    ws = wb.active
    rows = []
    for i in qs:
        v = getattr(i, "vehicle", None)
        rows.append([
            getattr(i, "created_at", "") and i.created_at.strftime("%Y-%m-%d %H:%M") or "",
            _vehicle_label(v) if v else "",
            getattr(i, "inspection_type", ""),
            getattr(i, "status", ""),
            getattr(i, "due_date", "") or "",
            getattr(i, "performed_on", "") or "",
            getattr(i, "notes", ""),
        ])

    _write_sheet(ws, "Inspections", ["Created At", "Vehicle", "Type", "Status", "Due Date", "Performed On", "Notes"], rows)
    return _xlsx_response(wb, "inspections.xlsx")


@login_required
def export_documents_xlsx(request):
    tenant = request.tenant
    qs = VehicleDocument.objects.filter(tenant=tenant).select_related("vehicle").order_by("-uploaded_at")

    wb = Workbook()
    ws = wb.active
    rows = []
    for d in qs:
        rows.append([
            d.uploaded_at.strftime("%Y-%m-%d %H:%M") if getattr(d, "uploaded_at", None) else "",
            _vehicle_label(d.vehicle),
            d.doc_type,
            d.title,
            d.expires_on or "",
            getattr(d.file, "url", ""),
        ])

    _write_sheet(ws, "Documents", ["Uploaded At", "Vehicle", "Doc Type", "Title", "Expires On", "File"], rows)
    return _xlsx_response(wb, "documents.xlsx")


@login_required
def export_inspection_alerts_xlsx(request):
    tenant = request.tenant
    qs = InspectionAlert.objects.filter(tenant=tenant).select_related("vehicle").order_by("-created_at")

    wb = Workbook()
    ws = wb.active
    rows = []
    for a in qs:
        v = getattr(a, "vehicle", None)
        rows.append([
            a.created_at.strftime("%Y-%m-%d %H:%M") if getattr(a, "created_at", None) else "",
            _vehicle_label(v) if v else "",
            getattr(a, "severity", ""),
            getattr(a, "status", ""),
            getattr(a, "title", ""),
            getattr(a, "detail", ""),
        ])

    _write_sheet(ws, "Inspection Alerts", ["Created At", "Vehicle", "Severity", "Status", "Title", "Detail"], rows)
    return _xlsx_response(wb, "inspection_alerts.xlsx")


# ---------------- CUSTOM REPORTS: WEEKLY + MONTHLY ----------------

def _range_from_query(request, default_days: int):
    today = timezone.localdate()
    start_s = (request.GET.get("start") or "").strip()
    end_s = (request.GET.get("end") or "").strip()

    if start_s and end_s:
        try:
            start = timezone.datetime.fromisoformat(start_s).date()
            end = timezone.datetime.fromisoformat(end_s).date()
            return start, end
        except Exception:
            pass

    end = today
    start = today - timedelta(days=default_days)
    return start, end


def _month_bounds(day):
    first = day.replace(day=1)
    # next month
    if first.month == 12:
        nxt = first.replace(year=first.year + 1, month=1)
    else:
        nxt = first.replace(month=first.month + 1)
    last = nxt - timedelta(days=1)
    return first, last


@login_required
def weekly_report(request):
    tenant = request.tenant
    start, end = _range_from_query(request, default_days=7)
    today = timezone.localdate()

    # KPIs within range
    insp_completed = (
        Inspection.objects
        .filter(
            tenant=tenant,
            inspection_date__isnull=False,
            inspection_date__gte=start,
            inspection_date__lte=end,
            status=Inspection.STATUS_COMPLETED,
        )
        .count()
    )

    alerts_created = (
        InspectionAlert.objects
        .filter(tenant=tenant, created_at__date__gte=start, created_at__date__lte=end)
        .count()
    )

    overdue_now = (
        Inspection.objects
        .filter(tenant=tenant, due_date__isnull=False, due_date__lt=today)
        .exclude(status=Inspection.STATUS_COMPLETED)
        .count()
    )

    docs_expiring_30 = (
        VehicleDocument.objects
        .filter(tenant=tenant, expires_on__isnull=False, expires_on__gte=today, expires_on__lte=today + timedelta(days=30))
        .count()
    )

    fuel_spend = (
        FuelLog.objects
        .filter(tenant=tenant, fuel_date__gte=start, fuel_date__lte=end)
        .exclude(cost__isnull=True)
        .aggregate(total=Coalesce(Sum("cost"), Decimal("0.00")))["total"]
    )

    # Charts
    fuel_daily = (
        FuelLog.objects
        .filter(tenant=tenant, fuel_date__gte=start, fuel_date__lte=end)
        .exclude(cost__isnull=True)
        .annotate(d=TruncDate("fuel_date"))
        .values("d")
        .annotate(total=Coalesce(Sum("cost"), Decimal("0.00")))
        .order_by("d")
    )
    fuel_labels = [r["d"].strftime("%Y-%m-%d") for r in fuel_daily]
    fuel_values = [float(r["total"]) for r in fuel_daily]

    alerts_daily = (
        InspectionAlert.objects
        .filter(tenant=tenant, created_at__date__gte=start, created_at__date__lte=end)
        .values("created_at__date")
        .annotate(total=Coalesce(Sum(Decimal("1.00")), Decimal("0.00")))
    )
    # normalize alerts per day (simple dict)
    ad = {}
    for r in alerts_daily:
        d = r["created_at__date"]
        ad[d.strftime("%Y-%m-%d")] = int(r["total"])
    alert_labels = fuel_labels[:] if fuel_labels else []
    if not alert_labels:
        # build date labels from range even if no fuel
        cur = start
        while cur <= end:
            alert_labels.append(cur.strftime("%Y-%m-%d"))
            cur += timedelta(days=1)
    alert_values = [ad.get(lbl, 0) for lbl in alert_labels]

    # Top vehicles by fuel spend
    top = (
        FuelLog.objects
        .filter(tenant=tenant, fuel_date__gte=start, fuel_date__lte=end)
        .exclude(cost__isnull=True)
        .values("vehicle_id")
        .annotate(total=Coalesce(Sum("cost"), Decimal("0.00")))
        .order_by("-total")[:10]
    )
    vehicle_map = {v.id: _vehicle_label(v) for v in Vehicle.objects.filter(tenant=tenant)}
    top_rows = [(vehicle_map.get(r["vehicle_id"], f"Vehicle #{r['vehicle_id']}"), float(r["total"])) for r in top]

    stale_list = vehicles_missing_fuel_logs(tenant, days=30)

    return render(request, "reports/weekly.html", {
        "start": start,
        "end": end,
        "insp_completed": insp_completed,
        "alerts_created": alerts_created,
        "overdue_now": overdue_now,
        "docs_expiring_30": docs_expiring_30,
        "fuel_spend": fuel_spend,
        "fuel_labels_json": json.dumps(fuel_labels),
        "fuel_values_json": json.dumps(fuel_values),
        "alert_labels_json": json.dumps(alert_labels),
        "alert_values_json": json.dumps(alert_values),
        "top_rows": top_rows,
        "stale_list": stale_list,
    })


@login_required
def monthly_report(request):
    tenant = request.tenant
    today = timezone.localdate()

    # Default: last full month
    first_this_month = today.replace(day=1)
    last_month_end = first_this_month - timedelta(days=1)
    start_default, end_default = _month_bounds(last_month_end)

    start_s = (request.GET.get("start") or "").strip()
    end_s = (request.GET.get("end") or "").strip()
    if start_s and end_s:
        try:
            start = timezone.datetime.fromisoformat(start_s).date()
            end = timezone.datetime.fromisoformat(end_s).date()
        except Exception:
            start, end = start_default, end_default
    else:
        start, end = start_default, end_default

    fuel_spend = (
        FuelLog.objects
        .filter(tenant=tenant, fuel_date__gte=start, fuel_date__lte=end)
        .exclude(cost__isnull=True)
        .aggregate(total=Coalesce(Sum("cost"), Decimal("0.00")))["total"]
    )

    # Compare with previous month
    prev_end = start - timedelta(days=1)
    prev_start, prev_end2 = _month_bounds(prev_end)

    prev_spend = (
        FuelLog.objects
        .filter(tenant=tenant, fuel_date__gte=prev_start, fuel_date__lte=prev_end2)
        .exclude(cost__isnull=True)
        .aggregate(total=Coalesce(Sum("cost"), Decimal("0.00")))["total"]
    )
    delta = float(fuel_spend) - float(prev_spend)

    # Charts: weekly within the range (group by week via trunc date buckets is messy)
    # We'll do daily line + top vehicles bar + alert severity pie (if field exists)
    daily = (
        FuelLog.objects
        .filter(tenant=tenant, fuel_date__gte=start, fuel_date__lte=end)
        .exclude(cost__isnull=True)
        .annotate(d=TruncDate("fuel_date"))
        .values("d")
        .annotate(total=Coalesce(Sum("cost"), Decimal("0.00")))
        .order_by("d")
    )
    daily_labels = [r["d"].strftime("%Y-%m-%d") for r in daily]
    daily_values = [float(r["total"]) for r in daily]

    top = (
        FuelLog.objects
        .filter(tenant=tenant, fuel_date__gte=start, fuel_date__lte=end)
        .exclude(cost__isnull=True)
        .values("vehicle_id")
        .annotate(total=Coalesce(Sum("cost"), Decimal("0.00")))
        .order_by("-total")[:10]
    )
    vehicle_map = {v.id: _vehicle_label(v) for v in Vehicle.objects.filter(tenant=tenant)}
    top_labels = [vehicle_map.get(r["vehicle_id"], f"Vehicle #{r['vehicle_id']}") for r in top]
    top_values = [float(r["total"]) for r in top]

    # Alerts by severity (works if your model has severity field; otherwise we fall back to status)
    sev_counts = {}
    if hasattr(InspectionAlert, "severity"):
        qs = (
            InspectionAlert.objects
            .filter(tenant=tenant, created_at__date__gte=start, created_at__date__lte=end)
            .values("severity")
            .annotate(total=Coalesce(Sum(Decimal("1.00")), Decimal("0.00")))
        )
        for r in qs:
            sev_counts[str(r["severity"])] = int(r["total"])
    else:
        qs = (
            InspectionAlert.objects
            .filter(tenant=tenant, created_at__date__gte=start, created_at__date__lte=end)
            .values("status")
            .annotate(total=Coalesce(Sum(Decimal("1.00")), Decimal("0.00")))
        )
        for r in qs:
            sev_counts[str(r["status"])] = int(r["total"])

    sev_labels = list(sev_counts.keys())
    sev_values = [sev_counts[k] for k in sev_labels]

    return render(request, "reports/monthly.html", {
        "start": start,
        "end": end,
        "fuel_spend": fuel_spend,
        "prev_spend": prev_spend,
        "delta": delta,
        "daily_labels_json": json.dumps(daily_labels),
        "daily_values_json": json.dumps(daily_values),
        "top_labels_json": json.dumps(top_labels),
        "top_values_json": json.dumps(top_values),
        "sev_labels_json": json.dumps(sev_labels),
        "sev_values_json": json.dumps(sev_values),
    })


@login_required
def export_weekly_xlsx(request):
    # Exports weekly report summary tables
    tenant = request.tenant
    start, end = _range_from_query(request, default_days=7)

    wb = Workbook()
    ws = wb.active

    fuel_qs = (
        FuelLog.objects
        .filter(tenant=tenant, fuel_date__gte=start, fuel_date__lte=end)
        .select_related("vehicle")
        .order_by("-fuel_date", "-created_at")
    )

    rows = []
    for r in fuel_qs:
        rows.append([
            r.fuel_date,
            _vehicle_label(r.vehicle),
            r.odometer or "",
            float(r.gallons),
            float(r.cost) if r.cost is not None else "",
            r.vendor,
            r.fuel_type,
        ])
    _write_sheet(ws, "Weekly Fuel", ["Fuel Date", "Vehicle", "Odometer", "Gallons", "Cost", "Vendor", "Fuel Type"], rows)

    return _xlsx_response(wb, f"weekly_report_{start}_{end}.xlsx")


@login_required
def export_monthly_xlsx(request):
    tenant = request.tenant
    # Default: last month
    today = timezone.localdate()
    first_this_month = today.replace(day=1)
    last_month_end = first_this_month - timedelta(days=1)
    start_default, end_default = _month_bounds(last_month_end)

    start_s = (request.GET.get("start") or "").strip()
    end_s = (request.GET.get("end") or "").strip()
    if start_s and end_s:
        try:
            start = timezone.datetime.fromisoformat(start_s).date()
            end = timezone.datetime.fromisoformat(end_s).date()
        except Exception:
            start, end = start_default, end_default
    else:
        start, end = start_default, end_default

    wb = Workbook()
    ws = wb.active

    fuel_qs = (
        FuelLog.objects
        .filter(tenant=tenant, fuel_date__gte=start, fuel_date__lte=end)
        .select_related("vehicle")
        .order_by("-fuel_date", "-created_at")
    )

    rows = []
    for r in fuel_qs:
        rows.append([
            r.fuel_date,
            _vehicle_label(r.vehicle),
            r.odometer or "",
            float(r.gallons),
            float(r.cost) if r.cost is not None else "",
            r.vendor,
            r.fuel_type,
        ])
    _write_sheet(ws, "Monthly Fuel", ["Fuel Date", "Vehicle", "Odometer", "Gallons", "Cost", "Vendor", "Fuel Type"], rows)

    return _xlsx_response(wb, f"monthly_report_{start}_{end}.xlsx")
